"""
Run dynasty model on sample CSV data → Excel rankings + JSON for dashboard
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
import numpy as np
import json
from dynasty_model import PlayerData, score_player, rank_players, to_json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


def load_csv(path: str) -> list[PlayerData]:
    df = pd.read_csv(path)
    players = []
    for _, row in df.iterrows():
        def g(col, default=None):
            val = row.get(col, default)
            if pd.isna(val):
                return default
            return val

        p = PlayerData(
            name=g("name", ""),
            position=g("position", ""),
            college=g("college", ""),
            draft_round=int(g("draft_round")) if g("draft_round") is not None else None,
            draft_pick=int(g("draft_pick")) if g("draft_pick") is not None else None,
            forty_time=g("forty_time"),
            height_inches=g("height_inches"),
            weight_lbs=g("weight_lbs"),
            vertical_inches=g("vertical_inches"),
            shuttle_seconds=g("shuttle_seconds"),
            three_cone=g("three_cone"),
            seasons=int(g("seasons", 1)),
            games=int(g("games", 0)),
            age_at_draft=g("age_at_draft"),
            completions=int(g("completions", 0)),
            attempts=int(g("attempts", 0)),
            pass_yards=int(g("pass_yards", 0)),
            pass_tds=int(g("pass_tds", 0)),
            interceptions=int(g("interceptions", 0)),
            years_as_starter=int(g("years_as_starter", 0)),
            rush_yards_qb=int(g("rush_yards_qb", 0)),
            qb_seasons=int(g("qb_seasons", 1)),
            rec_yards=int(g("rec_yards", 0)),
            receptions=int(g("receptions", 0)),
            rec_tds=int(g("rec_tds", 0)),
            rush_yards=int(g("rush_yards", 0)),
            rush_attempts=int(g("rush_attempts", 0)),
            rush_tds=int(g("rush_tds", 0)),
            targets=int(g("targets", 0)),
            team_pass_yards=int(g("team_pass_yards", 0)),
            team_pass_attempts=int(g("team_pass_attempts", 0)),
            team_rush_yards=int(g("team_rush_yards", 0)),
            team_total_tds=int(g("team_total_tds", 0)),
            breakout_age=g("breakout_age"),
        )
        players.append(p)
    return players


TIER_COLORS = {
    "Tier 1": "1A6B3C",
    "Tier 2": "2E86AB",
    "Tier 3": "F4A261",
    "Tier 4": "E76F51",
    "Tier 5": "9B2226",
}

POS_COLORS = {
    "QB": "264653",
    "RB": "2A9D8F",
    "WR": "E9C46A",
    "TE": "E76F51",
}


def build_excel(players: list[PlayerData], output_path: str):
    wb = Workbook()

    # ── Sheet 1: Overall Rankings ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Overall Rankings"

    header_fill = PatternFill("solid", start_color="0D1B2A", end_color="0D1B2A")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    results = [score_player(p) for p in players]
    results.sort(key=lambda x: -x["total_score"])

    cols = ["Rank", "Name", "Position", "College", "Score", "Tier", "Dynasty Value",
            "Draft Capital", "Dominator Rtg", "Speed Score", "Age Score", "Breakout Age"]
    ws.append(cols)

    for i, r in enumerate(results, 1):
        nm = r["normalized_metrics"]
        row = [
            i,
            r["name"],
            r["position"],
            r["college"],
            r["total_score"],
            r["tier"],
            r["dynasty_value"],
            round(nm.get("draft_capital", 0), 1),
            round(nm.get("dominator_rating", 0), 1),
            round(nm.get("speed_score", nm.get("height_adj_speed", 0)), 1),
            round(nm.get("age_at_draft", 0), 1),
            round(nm.get("breakout_age", 0), 1),
        ]
        ws.append(row)

    # Header formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Data formatting
    for row_idx in range(2, len(results) + 2):
        row = ws[row_idx]
        tier = ws.cell(row_idx, 6).value
        pos = ws.cell(row_idx, 3).value
        tier_color = TIER_COLORS.get(tier, "FFFFFF")
        pos_color = POS_COLORS.get(pos, "888888")

        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = center

        ws.cell(row_idx, 6).fill = PatternFill("solid", start_color=tier_color, end_color=tier_color)
        ws.cell(row_idx, 6).font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        ws.cell(row_idx, 3).fill = PatternFill("solid", start_color=pos_color, end_color=pos_color)
        ws.cell(row_idx, 3).font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

        # Alternate row shading
        if row_idx % 2 == 0:
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = PatternFill("solid", start_color="F7F9FC", end_color="F7F9FC")

    # Color scale on Score column
    ws.conditional_formatting.add(f"E2:E{len(results)+1}", ColorScaleRule(
        start_type="min", start_color="FF6B6B",
        mid_type="percentile", mid_value=50, mid_color="FFD93D",
        end_type="max", end_color="6BCB77"
    ))

    col_widths = [6, 22, 10, 18, 8, 10, 22, 13, 13, 12, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # ── Per-position sheets ────────────────────────────────────────────────────
    for pos in ["QB", "RB", "WR", "TE"]:
        pos_results = [r for r in results if r["position"] == pos]
        if not pos_results:
            continue
        ws2 = wb.create_sheet(f"{pos} Rankings")
        weight_keys = list(pos_results[0]["weights"].keys())
        headers = ["Rank", "Name", "College", "Score", "Tier", "Dynasty Value"] + \
                  [k.replace("_", " ").title() for k in weight_keys]
        ws2.append(headers)

        for i, r in enumerate(pos_results, 1):
            nm = r["normalized_metrics"]
            row = [i, r["name"], r["college"], r["total_score"], r["tier"], r["dynasty_value"]]
            row += [round(nm.get(k, 0), 1) for k in weight_keys]
            ws2.append(row)

        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx in range(2, len(pos_results) + 2):
            for cell in ws2[row_idx]:
                cell.font = body_font
                cell.alignment = center
                cell.border = border

        for col_idx in range(7, len(weight_keys) + 7):
            col_letter = get_column_letter(col_idx)
            ws2.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{len(pos_results)+1}",
                ColorScaleRule(start_type="min", start_color="FF6B6B",
                               mid_type="percentile", mid_value=50, mid_color="FFD93D",
                               end_type="max", end_color="6BCB77")
            )

        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = max(
                14, len(str(col[0].value or "")) + 2)
        ws2.freeze_panes = "A2"

    # ── Weights sheet ──────────────────────────────────────────────────────────
    from dynasty_model import WEIGHTS
    ws3 = wb.create_sheet("Metric Weights")
    ws3.append(["Position", "Metric", "Weight"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for pos2, wts in WEIGHTS.items():
        for metric, wt in wts.items():
            ws3.append([pos2, metric.replace("_", " ").title(), f"{wt*100:.0f}%"])

    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = center
            cell.border = border

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 10

    wb.save(output_path)
    print(f"✅ Excel saved: {output_path}")


def build_json_for_dashboard(players: list[PlayerData]) -> str:
    results = [score_player(p) for p in players]
    results.sort(key=lambda x: -x["total_score"])
    for i, r in enumerate(results, 1):
        r["rank"] = i
    return json.dumps(results, indent=2)


if __name__ == "__main__":
    import os
    base = os.path.dirname(__file__)
    csv_path = os.path.join(base, "sample_players.csv")
    xlsx_path = os.path.join(base, "dynasty_rankings.xlsx")
    json_path = os.path.join(base, "rankings_data.json")

    players = load_csv(csv_path)
    build_excel(players, xlsx_path)

    data = build_json_for_dashboard(players)
    with open(json_path, "w") as f:
        f.write(data)
    print(f"✅ JSON saved: {json_path}")

    # Print quick summary
    df = rank_players(players)
    print("\n📊 DYNASTY RANKINGS PREVIEW")
    print("=" * 65)
    print(df[["Rank", "Name", "Position", "Score", "Tier", "Dynasty Value"]].to_string(index=False))
